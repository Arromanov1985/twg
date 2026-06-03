from __future__ import annotations

from pathlib import Path
import re
from datetime import datetime

ROOT = Path(".")
APP = ROOT / "app.py"
KP = ROOT / "src" / "kp_generator.py"
TPL = ROOT / "templates" / "kp_template.html"
CATALOG = ROOT / "data" / "equipment_catalog.xlsx"


def backup_text(path: Path) -> None:
    if path.exists():
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path.with_name(path.name + f".bak_{stamp}").write_text(path.read_text(encoding="utf-8"), encoding="utf-8")


def sub_once(text: str, pattern: str, repl: str, name: str, flags=re.S) -> str:
    new, count = re.subn(pattern, repl, text, count=1, flags=flags)
    if count != 1:
        raise RuntimeError(f"Не удалось заменить блок: {name}")
    return new


if not APP.exists():
    raise SystemExit("app.py не найден. Запустите скрипт из корня репозитория twg.")

app = APP.read_text(encoding="utf-8")
backup_text(APP)

if "from datetime import date" not in app:
    app = app.replace("import operator\n", "import operator\nfrom datetime import date\n", 1)

price_helpers = '''
PRICE_COLUMN_ALIASES = {
    "retail": ["Розница", "Цена Розница", "Цена розница", "retail_price", "price_retail", "price"],
    "partner": ["Партнер", "Партнёр", "Цена Партнер", "Цена партнера", "Цена партнёра", "partner_price", "price_partner"],
}


def _first_existing_column(df: pd.DataFrame, names: list[str]) -> str | None:
    normalized = {str(col).strip().lower(): col for col in df.columns}
    for name in names:
        key = name.strip().lower()
        if key in normalized:
            return normalized[key]
    return None


def ensure_price_columns(catalog: pd.DataFrame) -> pd.DataFrame:
    catalog = catalog.copy()

    retail_col = _first_existing_column(catalog, PRICE_COLUMN_ALIASES["retail"])
    partner_col = _first_existing_column(catalog, PRICE_COLUMN_ALIASES["partner"])

    if retail_col is None:
        catalog["Розница"] = 0
        retail_col = "Розница"

    catalog["Розница"] = pd.to_numeric(catalog[retail_col], errors="coerce").fillna(0)

    if partner_col is None:
        catalog["Партнер"] = catalog["Розница"]
    else:
        catalog["Партнер"] = pd.to_numeric(catalog[partner_col], errors="coerce").fillna(catalog["Розница"])

    catalog["Выгода"] = catalog["Розница"] - catalog["Партнер"]

    # ВАЖНО: для клиента и КП используем только розничную цену.
    catalog["retail_price"] = catalog["Розница"]
    catalog["partner_price"] = catalog["Партнер"]
    catalog["benefit"] = catalog["Выгода"]
    catalog["price"] = catalog["Розница"]

    return catalog

'''

if "def ensure_price_columns(" not in app:
    app = app.replace("\n\ndef read_excel", "\n" + price_helpers + "\ndef read_excel", 1)

if "catalog = ensure_price_columns(catalog)" not in app:
    app = app.replace(
        '    catalog["price"] = pd.to_numeric(catalog["price"], errors="coerce").fillna(0)\n',
        '    catalog["price"] = pd.to_numeric(catalog["price"], errors="coerce").fillna(0)\n    catalog = ensure_price_columns(catalog)\n',
        1,
    )

new_build_input_form = '''def build_input_form(analysis: pd.DataFrame) -> dict[str, Any]:
    values: dict[str, Any] = {}
    st.subheader("Анализ воды и объект")

    meta_cols = st.columns(2)
    with meta_cols[0]:
        values["analysis_number"] = st.text_input("Номер анализа", "")
    with meta_cols[1]:
        analysis_date = st.date_input("Дата анализа", value=date.today())
        values["analysis_date"] = analysis_date.strftime("%d.%m.%Y")

    cols = st.columns(4)
    for i, row in analysis.iterrows():
        parameter = str(row["parameter"]).strip()
        label = str(row.get("label", parameter))
        unit = str(row.get("unit", ""))
        default = row.get("value", 0)
        with cols[i % 4]:
            if str(default).lower() in {"yes", "no", "true", "false", "да", "нет"}:
                values[parameter] = st.selectbox(
                    label,
                    ["нет", "да"],
                    index=1 if str(default).lower() in {"yes", "true", "да"} else 0,
                )
            else:
                values[parameter] = st.number_input(
                    f"{label}, {unit}" if unit else label,
                    value=float(default),
                    step=0.1,
                )
    return values
'''

app = sub_once(
    app,
    r'def build_input_form\(analysis: pd\.DataFrame\) -> dict\[str, Any\]:\n.*?\n\ndef build_odor_form',
    new_build_input_form + "\n\ndef build_odor_form",
    "build_input_form",
)

new_build_client_form = '''def build_client_form() -> dict[str, Any]:
    st.subheader("Данные объекта")
    o1, o2, o3 = st.columns(3)
    with o1:
        object_type = st.text_input("Тип объекта", "Частный дом")
    with o2:
        water_source = st.text_input("Источник воды", "Скважина")
    with o3:
        address = st.text_input("Адрес/объект для КП", "")

    st.subheader("Контакты")
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        company = st.text_input("Компания", "Частный клиент")
    with c2:
        manager = st.text_input("Менеджер", "TerraWater Group")
    with c3:
        phone = st.text_input("Телефон", "")
    with c4:
        email = st.text_input("Почта", "")

    return {
        "client_name": company,
        "company": company,
        "object_type": object_type,
        "water_source": water_source,
        "manager": manager,
        "phone": phone,
        "email": email,
        "address": address,
    }
'''

app = sub_once(
    app,
    r'def build_client_form\(\) -> dict\[str, Any\]:\n.*?\n\ndef export_kp_block',
    new_build_client_form + "\n\ndef export_kp_block",
    "build_client_form",
)

new_enrich = '''def enrich_analysis_for_kp(analysis: pd.DataFrame, values: dict[str, Any]) -> pd.DataFrame:
    extra_rows = [
        {
            "parameter": "analysis_number",
            "value": values.get("analysis_number", ""),
            "unit": "",
            "label": "Номер анализа",
        },
        {
            "parameter": "analysis_date",
            "value": values.get("analysis_date", ""),
            "unit": "",
            "label": "Дата анализа",
        },
    ]

    if "product_line" in values:
        extra_rows.append({
            "parameter": "product_line",
            "value": values.get("product_line", ""),
            "unit": "",
            "label": "Линейка загрузок",
        })

    extra_rows.extend([
        {
            "parameter": "odor_type",
            "value": values.get("odor_type", "Нет запаха"),
            "unit": "",
            "label": "Тип запаха",
        },
        {
            "parameter": "odor_level",
            "value": values.get("odor_level", "Нет"),
            "unit": "",
            "label": "Интенсивность запаха",
        },
    ])

    return pd.concat([analysis, pd.DataFrame(extra_rows)], ignore_index=True)
'''

app = sub_once(
    app,
    r'def enrich_analysis_for_kp\(analysis: pd\.DataFrame, values: dict\[str, Any\]\) -> pd\.DataFrame:\n.*?\n\ndef image_path',
    new_enrich + "\n\ndef image_path",
    "enrich_analysis_for_kp",
)

if "Внутренний расчет выгоды" not in app:
    app = app.replace(
        '    if reasons:\n        st.subheader("Почему выбрано это оборудование")',
        '''    if {"retail_price", "partner_price", "benefit"}.issubset(selected_df.columns):
        with st.expander("Внутренний расчет выгоды (не выводится в КП)"):
            internal_table = selected_df[["name", "code", "retail_price", "partner_price", "benefit"]].copy()
            internal_table.rename(
                columns={
                    "name": "Наименование",
                    "code": "Модель",
                    "retail_price": "Розница, ₽",
                    "partner_price": "Партнер, ₽",
                    "benefit": "Выгода, ₽",
                },
                inplace=True,
            )
            st.dataframe(internal_table, width="stretch", hide_index=True)
            st.info(f"Итого выгода: {selected_df['benefit'].sum():,.0f} ₽".replace(",", " "))

    if reasons:
        st.subheader("Почему выбрано это оборудование")''',
        1,
    )

APP.write_text(app, encoding="utf-8")

# kp_generator.py
if KP.exists():
    kp = KP.read_text(encoding="utf-8")
    backup_text(KP)

    if "def _row_price(" not in kp:
        kp = kp.replace(
            "\n\ndef _is_visible_in_kp",
            '''

def _row_price(row: pd.Series) -> float:
    for key in ("retail_price", "Розница", "price"):
        try:
            value = row.get(key, None)
            if value is not None and not pd.isna(value):
                return float(value)
        except Exception:
            continue
    return 0.0
''' + "\n\ndef _is_visible_in_kp",
            1,
        )

    kp = kp.replace('"price": float(row.get("price", 0) or 0),', '"price": _row_price(row),')
    kp = kp.replace('"price_text": _fmt_money(row.get("price", 0)),', '"price_text": _fmt_money(_row_price(row)),')
    KP.write_text(kp, encoding="utf-8")

# kp_template.html
if TPL.exists():
    tpl = TPL.read_text(encoding="utf-8")
    backup_text(TPL)

    tpl = tpl.replace(
        "Дата: {{ date }} · Клиент: {{ client.client_name }} · Менеджер: {{ client.manager }}",
        "Дата КП: {{ date }} · Клиент: {{ client.client_name }} · Менеджер: {{ client.manager }}",
    )

    if "values.analysis_number" not in tpl:
        tpl = tpl.replace(
            '<div>Дата КП: {{ date }} · Клиент: {{ client.client_name }} · Менеджер: {{ client.manager }}</div>',
            '''<div>Дата КП: {{ date }} · Клиент: {{ client.client_name }} · Менеджер: {{ client.manager }}</div>
      {% if values.analysis_number or values.analysis_date %}
      <div>Анализ воды{% if values.analysis_number %} № {{ values.analysis_number }}{% endif %}{% if values.analysis_date %} от {{ values.analysis_date }}{% endif %}</div>
      {% endif %}''',
            1,
        )

    if "client.phone" not in tpl:
        tpl = tpl.replace(
            '{% if client.address %}<p>📍 {{ client.address }}</p>{% endif %}',
            '''{% if client.address %}<p>📍 {{ client.address }}</p>{% endif %}
      {% if client.phone %}<p>☎ {{ client.phone }}</p>{% endif %}
      {% if client.email %}<p>✉ {{ client.email }}</p>{% endif %}''',
            1,
        )

        tpl = tpl.replace(
            "<div>8 800 333-02-42<br>www.terrawater.ru<br>info@terrawater.ru</div>",
            '''<div><b>Контакты</b><br>{{ client.manager }}{% if client.phone %}<br>{{ client.phone }}{% endif %}{% if client.email %}<br>{{ client.email }}{% endif %}</div>''',
            1,
        )

    TPL.write_text(tpl, encoding="utf-8")

# equipment_catalog.xlsx
if CATALOG.exists():
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        wb = load_workbook(CATALOG)
        ws = wb.active
        headers = {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], start=1) if cell.value is not None}

        def ensure_col(name: str) -> int:
            if name in headers:
                return headers[name]
            col = ws.max_column + 1
            ws.cell(row=1, column=col, value=name)
            headers[name] = col
            return col

        retail_col = ensure_col("Розница")
        partner_col = ensure_col("Партнер")
        benefit_col = ensure_col("Выгода")

        source_price_col = headers.get("price") or headers.get("Цена") or retail_col

        for row in range(2, ws.max_row + 1):
            retail = ws.cell(row=row, column=retail_col).value
            if retail in (None, ""):
                ws.cell(row=row, column=retail_col, value=ws.cell(row=row, column=source_price_col).value or 0)

            partner = ws.cell(row=row, column=partner_col).value
            if partner in (None, ""):
                ws.cell(row=row, column=partner_col, value=ws.cell(row=row, column=retail_col).value or 0)

            ws.cell(row=row, column=benefit_col, value=f"={get_column_letter(retail_col)}{row}-{get_column_letter(partner_col)}{row}")

        for name in ("Розница", "Партнер", "Выгода"):
            ws.column_dimensions[get_column_letter(headers[name])].width = 14

        wb.save(CATALOG)
        print("Прайс обновлен: добавлены/проверены колонки Розница, Партнер, Выгода.")
    except Exception as exc:
        print("Прайс не удалось обновить автоматически, но приложение уже поддерживает Розница/Партнер/Выгода:", exc)

print("Готово.")
print("Изменены файлы: app.py, src/kp_generator.py, templates/kp_template.html")
print("Если прайс найден, также обновлен data/equipment_catalog.xlsx")
